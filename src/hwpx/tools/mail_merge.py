# SPDX-License-Identifier: Apache-2.0
"""Batch document generation helpers for HWPX templates."""

from __future__ import annotations

import csv
import json
import re
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from ..document import HwpxDocument
from .exporter import export_text
from .package_validator import validate_editor_open_safety
from .pii import DEFAULT_POLICY, PIIPolicy, mask_pii

MAIL_MERGE_REPORT_VERSION = "mail-merge-v1"

_PLACEHOLDER_RE = re.compile(
    r"(?P<brace>\{\{\s*(?P<brace_key>[A-Za-z_][A-Za-z0-9_.-]*)\s*\}\})"
    r"|(?P<dollar>\$\{\s*(?P<dollar_key>[A-Za-z_][A-Za-z0-9_.-]*)\s*\})"
    r"|(?P<angle><<\s*(?P<angle_key>[A-Za-z_][A-Za-z0-9_.-]*)\s*>>)"
)
_SAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9가-힣._ -]+")


def load_mail_merge_rows(data: str | Path | Sequence[Mapping[str, Any]] | Mapping[str, Any]) -> list[dict[str, Any]]:
    """Load mail-merge rows from CSV, JSON, or in-memory mappings."""

    if isinstance(data, (str, Path)):
        path = Path(data)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                return [dict(row) for row in csv.DictReader(stream)]
        if suffix == ".json":
            with path.open("r", encoding="utf-8") as stream:
                payload = json.load(stream)
            return load_mail_merge_rows(payload)
        if suffix in (".xlsx", ".xlsm"):
            return _load_xlsx_rows(path)
        raise ValueError(f"unsupported mail merge data file: {path}")

    if isinstance(data, Mapping):
        rows = data.get("rows", data.get("data"))
        if rows is None:
            raise ValueError("mail merge data mapping must contain rows or data")
        return load_mail_merge_rows(rows)  # type: ignore[arg-type]

    rows: list[dict[str, Any]] = []
    for index, row in enumerate(data):
        if not isinstance(row, Mapping):
            raise TypeError(f"mail merge row {index} must be a mapping")
        rows.append({str(key): value for key, value in row.items()})
    return rows


def _load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    """Read a roster (명부) from the first sheet of an .xlsx/.xlsm workbook.

    The first non-empty row is the header (placeholder keys); each later row is a
    record. Cells are coerced to ``str`` (so ``{{no}}`` from an integer cell merges
    cleanly) and fully-empty rows are dropped. ``openpyxl`` is an optional dependency
    (``python-hwpx[xlsx]``) — absent, a clear ImportError tells the caller to install.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise ImportError(
            "reading .xlsx rosters requires openpyxl — install python-hwpx[xlsx]"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header: list[str] | None = None
        rows: list[dict[str, Any]] = []
        for raw in sheet.iter_rows(values_only=True):
            if raw is None or all(cell is None for cell in raw):
                continue
            if header is None:
                header = [str(cell).strip() if cell is not None else "" for cell in raw]
                continue
            record: dict[str, Any] = {}
            for key, cell in zip(header, raw):
                if not key:
                    continue
                record[key] = "" if cell is None else str(cell)
            rows.append(record)
    finally:
        workbook.close()
    return rows


def inspect_mail_merge_placeholders(source: str | Path | HwpxDocument) -> dict[str, Any]:
    """Return placeholders found in a template document."""

    close_doc = False
    if isinstance(source, HwpxDocument):
        document = source
    else:
        document = HwpxDocument.open(source)
        close_doc = True
    try:
        text = export_text(document)
    finally:
        if close_doc:
            document.close()
    placeholders = _find_placeholders(text)
    keys = sorted({item["key"] for item in placeholders})
    return {
        "report_version": MAIL_MERGE_REPORT_VERSION,
        "placeholderCount": len(placeholders),
        "keys": keys,
        "placeholders": placeholders,
    }


def _iter_cells(document: HwpxDocument) -> list[Any]:
    """Every table cell in *document*, descending into nested tables."""

    cells: list[Any] = []

    def _walk_para(paragraph: Any) -> None:
        for table in getattr(paragraph, "tables", None) or []:
            for row in getattr(table, "rows", None) or []:
                for cell in getattr(row, "cells", None) or []:
                    cells.append(cell)
                    for cell_para in getattr(cell, "paragraphs", None) or []:
                        _walk_para(cell_para)

    for paragraph in document.paragraphs:
        _walk_para(paragraph)
    return cells


def _measure_placeholder_slots(
    document: HwpxDocument,
    placeholders: Sequence[Mapping[str, Any]],
    *,
    max_lines: int,
) -> dict[str, Any]:
    """Template-once-measure: each placeholder token -> its *narrowest* containing
    cell's :class:`SlotMetrics` (advance-model, no oracle). Tokens that live outside a
    measurable cell (free body text) get no slot, so their fit stays *unverified*
    (never a hard fail) — the same honesty as ``fill_form_field`` with no box_width.
    """

    from ..form_fit import resolve_slot_metrics

    cells = _iter_cells(document)
    slots: dict[str, Any] = {}
    for placeholder in placeholders:
        token = str(placeholder["token"])
        best = None
        for cell in cells:
            try:
                cell_text = cell.text
            except Exception:  # pragma: no cover - defensive on odd cells
                continue
            if token in cell_text:
                metrics = resolve_slot_metrics(cell, document, max_lines=max_lines)
                if best is None or metrics.available_width < best.available_width:
                    best = metrics
        if best is not None:
            slots[token] = best
    return slots


def mail_merge(
    template: str | Path,
    data: str | Path | Sequence[Mapping[str, Any]] | Mapping[str, Any],
    *,
    output_dir: str | Path | None = None,
    filename_pattern: str = "{index:03d}.hwpx",
    zip_path: str | Path | None = None,
    strict: bool = False,
    split_newlines: bool = True,
    fit_policy: "Any | None" = None,
    max_lines: int = 1,
    masking_policy: "PIIPolicy | None" = DEFAULT_POLICY,
) -> dict[str, Any]:
    """Generate one HWPX per row from a placeholder template.

    Supported placeholder forms are ``{{field}}``, ``${field}``, and
    ``<<field>>``. Missing values are reported per row; by default documents are
    still generated so an operator can inspect the partial output. Set
    ``strict=True`` to skip rows with missing required data.

    When *fit_policy* (a :class:`~hwpx.form_fit.policy.FitPolicy`) is given the batch
    is **fit-aware**: each placeholder's slot is measured **once** from the template
    (advance-model, no per-record oracle) and every record's value is fit against it.
    A value that would overflow its slot, or a row missing a required field, is
    isolated into ``needsReview[]`` / ``skipped[]`` with a reason — never silently
    truncated, never corrupting the rest of the batch (FR-004).
    """

    template_path = Path(template)
    rows = load_mail_merge_rows(data)
    if not rows:
        raise ValueError("mail merge requires at least one data row")

    output_root = Path(output_dir) if output_dir is not None else template_path.with_suffix("").parent / "mail_merge_out"
    output_root.mkdir(parents=True, exist_ok=True)

    placeholder_report = inspect_mail_merge_placeholders(template_path)
    placeholders = list(placeholder_report["placeholders"])
    required_keys = sorted({str(item["key"]) for item in placeholders})

    # template-once-measure: one slot measurement reused across every record.
    slots: dict[str, Any] = {}
    if fit_policy is not None:
        measure_doc = HwpxDocument.open(template_path)
        try:
            slots = _measure_placeholder_slots(measure_doc, placeholders, max_lines=max_lines)
        finally:
            measure_doc.close()

    row_reports: list[dict[str, Any]] = []
    generated_paths: list[Path] = []
    for index, row in enumerate(rows, start=1):
        missing_keys = [key for key in required_keys if _is_missing(row.get(key))]
        filename = _format_filename(filename_pattern, row=row, index=index)
        out_path = output_root / filename
        if out_path.suffix.lower() != ".hwpx":
            out_path = out_path.with_suffix(".hwpx")

        if strict and missing_keys:
            row_reports.append(
                {
                    "rowIndex": index,
                    "filename": str(out_path),
                    "created": False,
                    "replacedCount": 0,
                    "missingKeys": missing_keys,
                    "unresolvedPlaceholders": [item["token"] for item in placeholders if item["key"] in missing_keys],
                    "openSafety": None,
                    "fitFields": [],
                    "maskedFields": [],
                    "reasons": ["missing_required"],
                    "ok": False,
                }
            )
            continue

        document = HwpxDocument.open(template_path)
        row_cells = _iter_cells(document)
        replaced_count = 0
        fit_fields: list[dict[str, Any]] = []
        masked_fields: list[str] = []
        try:
            for placeholder in placeholders:
                key = str(placeholder["key"])
                token = str(placeholder["token"])
                value = "" if _is_missing(row.get(key)) else str(row.get(key))
                if masking_policy is not None:
                    masked_value = mask_pii(value, masking_policy)
                    if masked_value != value and key not in masked_fields:
                        masked_fields.append(key)
                    value = masked_value
                if not split_newlines:
                    value = value.replace("\r\n", " ").replace("\n", " ")
                apply_value = value
                if fit_policy is not None and value and token in slots:
                    fit_result = _fit_value(value, slots[token], fit_policy, field_id=key)
                    apply_value = fit_result.applied_value
                    if (not fit_result.ok) or fit_result.overflow_detected:
                        fit_fields.append(_fit_field_report(key, fit_result))
                replaced_count += _replace_token(document, token, apply_value, row_cells)
            document.save_to_path(out_path)
        finally:
            document.close()

        open_safety = validate_editor_open_safety(out_path).to_dict()
        unresolved = _unresolved_placeholders(out_path)
        reasons: list[str] = []
        if missing_keys:
            reasons.append("missing_required")
        if fit_fields:
            reasons.append("overflow")
        if unresolved:
            reasons.append("unresolved_placeholder")
        if not bool(open_safety["ok"]):
            reasons.append("open_safety")
        row_ok = not reasons
        row_reports.append(
            {
                "rowIndex": index,
                "filename": str(out_path),
                "created": True,
                "replacedCount": replaced_count,
                "missingKeys": missing_keys,
                "unresolvedPlaceholders": unresolved,
                "openSafety": open_safety,
                "fitFields": fit_fields,
                "maskedFields": masked_fields,
                "reasons": reasons,
                "ok": row_ok,
            }
        )
        generated_paths.append(out_path)

    zip_result: dict[str, Any] | None = None
    if zip_path is not None:
        zip_file = Path(zip_path)
        zip_file.parent.mkdir(parents=True, exist_ok=True)
        with ZipFile(zip_file, "w", compression=ZIP_DEFLATED) as archive:
            for path in generated_paths:
                archive.write(path, arcname=path.name)
        zip_result = {
            "path": str(zip_file),
            "entryCount": len(generated_paths),
            "entries": [path.name for path in generated_paths],
        }

    rows_with_issues = [
        report["rowIndex"] for report in row_reports if report.get("reasons")
    ]
    # per-record isolation (FR-004): skipped = never generated; needsReview = generated
    # but flagged. Reasons let an operator act without scanning the whole batch.
    skipped = [
        {
            "rowIndex": report["rowIndex"],
            "reasons": report.get("reasons", []),
            "missingKeys": report.get("missingKeys", []),
        }
        for report in row_reports
        if not report["created"]
    ]
    needs_review = [
        {
            "rowIndex": report["rowIndex"],
            "filename": report["filename"],
            "reasons": report.get("reasons", []),
            "fitFields": report.get("fitFields", []),
            "missingKeys": report.get("missingKeys", []),
            "unresolvedPlaceholders": report.get("unresolvedPlaceholders", []),
        }
        for report in row_reports
        if report["created"] and report.get("reasons")
    ]
    open_safety = _open_safety_summary(row_reports, created_count=len(generated_paths))
    return {
        "report_version": MAIL_MERGE_REPORT_VERSION,
        "template": str(template_path),
        "outputDir": str(output_root),
        "filenamePattern": filename_pattern,
        "placeholderKeys": required_keys,
        "fitAware": fit_policy is not None,
        "measuredSlots": sorted(slots.keys()),
        "rowCount": len(rows),
        "createdCount": len(generated_paths),
        "rowsWithIssues": rows_with_issues,
        "needsReview": needs_review,
        "skipped": skipped,
        "ok": not rows_with_issues and len(generated_paths) == len(rows),
        "openSafety": open_safety,
        "verification": {
            "openSafety": open_safety,
            "createdCount": len(generated_paths),
            "rowCount": len(rows),
            "rowsWithIssues": rows_with_issues,
            "zip": zip_result,
        },
        "rows": row_reports,
        "zip": zip_result,
    }


def _find_placeholders(text: str) -> list[dict[str, str]]:
    found: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for match in _PLACEHOLDER_RE.finditer(text):
        key = match.group("brace_key") or match.group("dollar_key") or match.group("angle_key")
        token = match.group(0)
        if not key or (token, key) in seen:
            continue
        seen.add((token, key))
        found.append({"token": token, "key": key})
    return found


def _unresolved_placeholders(path: Path) -> list[str]:
    document = HwpxDocument.open(path)
    try:
        text = export_text(document)
    finally:
        document.close()
    return [item["token"] for item in _find_placeholders(text)]


def _open_safety_summary(row_reports: Sequence[Mapping[str, Any]], *, created_count: int) -> dict[str, Any]:
    checked = 0
    failures: list[dict[str, Any]] = []
    for row in row_reports:
        open_safety = row.get("openSafety")
        if not isinstance(open_safety, Mapping):
            continue
        checked += 1
        if not bool(open_safety.get("ok")):
            failures.append(
                {
                    "rowIndex": row.get("rowIndex"),
                    "filename": row.get("filename"),
                    "summary": open_safety.get("summary"),
                }
            )
    return {
        "ok": checked == created_count and not failures,
        "checkedCount": checked,
        "failureCount": len(failures),
        "failures": failures,
    }


def _replace_token(document: HwpxDocument, token: str, value: str, cells: Sequence[Any]) -> int:
    """Replace *token* in body runs **and** table-cell runs.

    ``replace_text_in_runs`` only reaches body runs (``find_runs_by_style`` does not
    descend into ``<hp:tbl>`` cells), so a placeholder living in a 발신·결재/안내 표
    cell would silently survive as an unresolved token. Cells are handled per-run here
    (same per-run semantics as the body path).
    """

    count = document.replace_text_in_runs(token, value)
    for cell in cells:
        for cell_para in getattr(cell, "paragraphs", None) or []:
            for run in cell_para.runs:
                count += run.replace_text(token, value)
    return count


def _fit_value(value: str, slot: Any, fit_policy: Any, *, field_id: str) -> Any:
    """Fit *value* into *slot* under *fit_policy* (advance-model FitEngine)."""

    from ..form_fit import FitEngine

    return FitEngine().fit(value, slot, fit_policy, field_id=field_id)


def _fit_field_report(key: str, fit_result: Any) -> dict[str, Any]:
    """Compact per-field overflow record for ``needsReview`` (with retry advice)."""

    return {
        "key": key,
        "overflowDetected": bool(getattr(fit_result, "overflow_detected", False)),
        "confidence": getattr(fit_result, "confidence", "high"),
        "appliedValue": getattr(fit_result, "applied_value", None),
        "advice": fit_result.suggested_retry(),
    }


def _is_missing(value: Any) -> bool:
    return value is None or str(value) == ""


def _format_filename(pattern: str, *, row: Mapping[str, Any], index: int) -> str:
    values = _FormatValues(row, index=index)
    try:
        filename = pattern.format_map(values)
    except (KeyError, ValueError):
        filename = f"{index:03d}.hwpx"
    name = Path(str(filename)).name.strip() or f"{index:03d}.hwpx"
    return _SAFE_FILENAME_RE.sub("_", name)


class _FormatValues(dict[str, Any]):
    def __init__(self, row: Mapping[str, Any], *, index: int) -> None:
        super().__init__({str(key): value for key, value in row.items()})
        self["index"] = index
        self["index0"] = index - 1

    def __missing__(self, key: str) -> str:
        return ""


__all__ = [
    "MAIL_MERGE_REPORT_VERSION",
    "inspect_mail_merge_placeholders",
    "load_mail_merge_rows",
    "mail_merge",
]
